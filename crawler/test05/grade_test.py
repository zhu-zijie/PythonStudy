import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment

alignment = Alignment(horizontal='center', vertical='center')
side = Side(style='mediumDashed', color='ffcccc')
# open the workbook
wb = openpyxl.load_workbook('grade.xlsx')
sheet = wb.worksheets[0]

sheet.row_dimensions[1].height = 30
sheet.column_dimensions['E'].width = 120

sheet['E1'] = 'Average'
sheet['E5'].font = Font(bold=True, color='FF0000', size=18, name='Arial')
sheet.cell(1, 5).alignment = alignment
sheet.cell(1, 5).border = Border(top=side, bottom=side, left=side, right=side)

for index in range(2, 7):
    sheet[f'E{index}'] = f'=average(B{index}:D{index})'
    sheet[f'E{index}'].font = Font(bold=True, color='FF0000', size=18, name='Arial')
    sheet[f'E{index}'].border = Border(top=side, bottom=side, left=side, right=side)
    sheet[f'E{index}'].alignment = alignment

wb.save('grade.xlsx')
